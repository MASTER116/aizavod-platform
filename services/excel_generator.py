"""Генерация Excel-файлов для грантовых калькуляций."""
from __future__ import annotations

import json
import logging
import os
from datetime import datetime

logger = logging.getLogger("aizavod.excel_generator")


def generate_budget_excel(budget_json: str, output_dir: str = "/tmp") -> str:
    """Сгенерировать Excel из JSON бюджета. Возвращает путь к файлу."""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        logger.error("openpyxl не установлен")
        return ""

    # Очистка markdown-обёртки ```json ... ```
    cleaned = budget_json.strip()
    if cleaned.startswith("```"):
        # Убираем первую строку (```json) и последнюю (```)
        lines = cleaned.split("\n")
        lines = [l for l in lines if not l.strip().startswith("```")]
        cleaned = "\n".join(lines)

    try:
        data = json.loads(cleaned)
    except (json.JSONDecodeError, TypeError):
        logger.error("Невалидный JSON бюджета: %s", cleaned[:200])
        return ""

    wb = openpyxl.Workbook()

    # ─── Лист 1: Смета ─────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Смета проекта"

    # Стили
    header_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=11)
    section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    total_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Заголовок
    title = data.get("project_title", "Проект")
    ws.merge_cells("A1:F1")
    ws["A1"] = f"СМЕТА ПРОЕКТА: {title}"
    ws["A1"].font = header_font

    ws["A2"] = "Сумма гранта:"
    ws["B2"] = data.get("grant_amount", 0)
    ws["B2"].number_format = '#,##0 ₽'
    ws["A3"] = "Собственные средства:"
    ws["B3"] = data.get("own_funds", 0)
    ws["B3"].number_format = '#,##0 ₽'
    ws["A4"] = "Срок реализации (мес):"
    ws["B4"] = data.get("duration_months", 0)

    # Заголовки таблицы
    row = 6
    headers = ["№", "Статья расходов", "Ед. изм.", "Кол-во", "Цена (руб)", "Итого (руб)"]
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 18

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = section_font
        cell.fill = section_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    row += 1
    item_num = 0
    grand_total = 0

    for cat in data.get("categories", []):
        # Секция
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell = ws.cell(row=row, column=1, value=cat.get("name", ""))
        cell.font = section_font
        cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        cell.border = thin_border
        row += 1

        cat_total = 0
        for item in cat.get("items", []):
            item_num += 1
            values = [
                item_num,
                item.get("name", ""),
                item.get("unit", "шт"),
                item.get("quantity", 0),
                item.get("price", 0),
                item.get("total", 0),
            ]
            for col, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=v)
                cell.border = thin_border
                if col >= 4:
                    cell.number_format = '#,##0'
                if col == 6:
                    cell.number_format = '#,##0 ₽'
            cat_total += item.get("total", 0)
            row += 1

        # Подитог секции
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        cell = ws.cell(row=row, column=1, value=f"Итого {cat.get('name', '')}:")
        cell.font = section_font
        cell.border = thin_border
        cell = ws.cell(row=row, column=6, value=cat_total)
        cell.font = section_font
        cell.number_format = '#,##0 ₽'
        cell.border = thin_border
        grand_total += cat_total
        row += 1

    # Общий итог
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    cell = ws.cell(row=row, column=1, value="ИТОГО ПО ПРОЕКТУ:")
    cell.font = Font(bold=True, size=12)
    cell.fill = total_fill
    cell.border = thin_border
    cell = ws.cell(row=row, column=6, value=grand_total)
    cell.font = Font(bold=True, size=12)
    cell.number_format = '#,##0 ₽'
    cell.fill = total_fill
    cell.border = thin_border

    # ─── Лист 2: План работ ─────────────────────────────────────────────
    ws2 = wb.create_sheet("План работ")
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 15
    ws2.column_dimensions["C"].width = 40
    ws2.column_dimensions["D"].width = 18

    ws2["A1"] = "ПЛАН РАБОТ"
    ws2["A1"].font = header_font

    headers2 = ["Этап", "Период", "Результаты", "Бюджет (руб)"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.font = section_font
        cell.fill = section_fill
        cell.border = thin_border

    for i, stage in enumerate(data.get("timeline", []), 4):
        ws2.cell(row=i, column=1, value=stage.get("stage", "")).border = thin_border
        ws2.cell(row=i, column=2, value=stage.get("months", "")).border = thin_border
        deliverables = ", ".join(stage.get("deliverables", []))
        ws2.cell(row=i, column=3, value=deliverables).border = thin_border
        cell = ws2.cell(row=i, column=4, value=stage.get("budget", 0))
        cell.number_format = '#,##0 ₽'
        cell.border = thin_border

    # Сохранение
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    safe_title = "".join(c for c in title[:30] if c.isalnum() or c in " _-").strip().replace(" ", "_")
    filename = f"budget_{safe_title}_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)

    os.makedirs(output_dir, exist_ok=True)
    wb.save(filepath)
    logger.info("Excel сохранён: %s", filepath)
    return filepath
